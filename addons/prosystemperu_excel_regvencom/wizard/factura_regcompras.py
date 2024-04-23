# -*- coding: utf-8 -*-

from odoo import api, fields, models
from datetime import datetime
from odoo.exceptions import UserError, ValidationError

#sirve para generar el archivo excel
#import xlwt
import openpyxl
import base64
import io

class WizardFacturaRegCompras(models.TransientModel):
    _name = "factura.regcompras"    

    #funcion para capturar los años de facturas validadas en una lista que su contenido son tuplas
    def periodo_buscar(self):
       
        self.env.cr.execute("""SELECT distinct extract(year from date_invoice)::text || lpad(extract(month 
                                from date_invoice)::text,2,'0') from account_invoice order by 1 desc""")
        select_anno= self.env.cr.fetchall()         

        datito=""

        anho=[]         
        
        for line in select_anno:                
            for lineas in line:
                if lineas:
                    contenido_anho=[]
                    primero=""
                    segundo=""

                    primero=lineas
                    segundo= self.get_mes(lineas[4::]) + ' ' + lineas[0:4]

                    contenido_anho.append(primero)
                    contenido_anho.append(segundo)                                  
                    anho.append(tuple(contenido_anho))

        return anho
    #termina funcion para capturar los años de facturas validadas en una lista que su contenido son tuplas
    def get_mes(self,valor):
        if valor == '01':
            return 'ENERO'
        if valor == '02':
            return 'FEBRERO'
        if valor == '03':
            return 'MARZO'
        if valor == '04':
            return 'ABRIL'
        if valor == '05':
            return 'MAYO'
        if valor == '06':
            return 'JUNIO'
        if valor == '07':
            return 'JULIO'
        if valor == '08':
            return 'AGOSTO'
        if valor == '09':
            return 'SEPTIEMBRE'
        if valor == '10':
            return 'OCTUBRE'
        if valor == '11':
            return 'NOVIEMBRE'
        if valor == '12':
            return 'DICIEMBRE'


    #campos en el wizard
    periodo_consulta=fields.Selection(periodo_buscar)
    #mes_consulta = fields.Selection([('1','Enero'),('2','Febrero'),('3','Marzo'),('4','Abril'),('5','Mayo'),('6','Junio'),('7','Julio'),('8','Agosto'),('9','Septiembre'),('10','Octubre'),('11','Noviembre'),('12','Diciembre')])
    
    
    #funcion para que llame al reporte
    @api.multi
    def action_excel(self):

        wb = openpyxl.Workbook()
        ws = wb.active
        ws = wb['Sheet']

        ws['A1'] = 'FEC. EMISION'
        ws['B1'] = 'TD'
        ws['C1'] = 'SERIE'
        ws['D1'] = 'NUMERO'
        ws['E1'] = 'RUC'
        ws['F1'] = 'RAZON SOCIAL'
        ws['G1'] = 'MONEDA'
        ws['H1'] = 'SUBTOTAL'
        ws['I1'] = 'IGV'
        ws['J1'] = 'INAFECTA'
        ws['K1'] = 'EXONERADO'
        ws['L1'] = 'EXPO'
        ws['M1'] = 'PERC'
        ws['N1'] = 'TOTAL'
        ws['O1'] = 'T/C'
        ws['P1'] = 'TIPO DOC REF'
        ws['Q1'] = 'FECHA EMI REF'
        ws['R1'] = 'SERIE REF'
        ws['S1'] = 'NUMERO REF'


        #rango de celdas para poner estilo de la celda
        seleccion = ws['A1':'S1']
        dicc_cotiz = {}
        for filas in seleccion:
            for columnas in filas:
                columnas.fill = openpyxl.styles.PatternFill(fgColor="FF0000", fill_type = "solid")  #para el colo de fondo tambien se puede poner asi fgColor=openpyxl.styles.colors.RED
                #columnas.font = openpyxl.styles.Font(name='Calibri', size=11, bold=False,color=openpyxl.styles.colors.BLACK)  #esta linea es para estilo font de la celda
        #termina rango de celdas para poner estilo colores al fondo de la celda

        self.env.cr.execute((""" SELECT a.date_invoice as fec_emi, d.abreviatura td, a.serie_compra serie, a.numeracion_compra numero, 
                            coalesce(c.vat,'')::varchar ruc, 
                            coalesce(c.name,'')::varchar raz_social,
                            case when b.name = 'PEN' then 'L' else 'E' end mon, 
                            coalesce(a.total_operaciones_gravadas,0.00)::numeric(10,2) grav, 
                            coalesce(a.amount_tax,0.00)::numeric(10,2) igv, 
                            coalesce(a.total_operaciones_exoneradas,0.00)::numeric(10,2) exo,
                            coalesce(a.total_operaciones_inafectas,0.00)::numeric(10,2) ina, 
                            coalesce(a.total_operaciones_exportadas, 0.00)::numeric(10,2) exp,
                            ''::varchar m_percepcion, a.amount_total, 0 tc,
                            ''::varchar tipodoc_afecta, '01/01/1900'::text fecdoc_afecta,
                            ''::varchar seriedoc_afecta , ''::varchar numdoc_afecta
                            from account_invoice a 
                            left join res_currency b on(b.id = a.currency_id)
                            left join res_partner c on(c.id = a.partner_id)
                            left join tipo_documento d on(d.id = a.tip_doc_compra)
                            where extract(year from a.date_invoice)::text || lpad(extract(month from a.date_invoice)::text,2,'0') ='%s' and a.state in('open','paid') and substring(a.type, 1,3) ='out'
                            order by a.date_invoice, a.tipo_doc, a.serie, a.numeracion; """)% self.periodo_consulta)
        oconsulta = self.env.cr.fetchall()

        i=2
        for item in oconsulta:
            date_object = datetime.strptime(item[0], '%Y-%m-%d')
            # pasando el valor a la celda 
            ws.cell(i,1).value = date_object                      
            ws.cell(i,2).value = item[1]
            ws.cell(i,3).value = item[2]
            ws.cell(i,4).value = item[3]
            ws.cell(i,5).value = item[4]
            ws.cell(i,6).value = item[5]
            ws.cell(i,7).value = item[6]
            ws.cell(i,8).value = item[7]
            ws.cell(i,9).value = item[8]
            ws.cell(i,10).value = item[9]
            ws.cell(i,11).value = item[10]
            ws.cell(i,12).value = item[11]
            ws.cell(i,13).value = item[12]
            ws.cell(i,14).value = item[13]
            ws.cell(i,15).value = item[14]
            ws.cell(i,16).value = item[15]

            if  item[16]=='01/01/1900':
                ws.cell(i,17).value = item[16]

            else:
                date_object2 = datetime.strptime(item[16], '%Y-%m-%d')
                ws.cell(i,17).value = date_object2          #da el valor a ese celda
                ws.cell(i,17).number_format = 'dd/mm/yyyy'  #da el formato a esa celda

            ws.cell(i,18).value = item[17]
            ws.cell(i,19).value = item[18]

            #dando formato a cada celda
            ws.cell(i,1).number_format = 'dd/mm/yyyy'            
            ws.cell(i,8).number_format = '#,##0.00'
            ws.cell(i,9).number_format = '#,##0.00'
            ws.cell(i,10).number_format = '#,##0.00'
            ws.cell(i,11).number_format = '#,##0.00'
            ws.cell(i,12).number_format = '#,##0.00'
            ws.cell(i,13).number_format = '#,##0.00'
            ws.cell(i,14).number_format = '#,##0.00'
            ws.cell(i,15).number_format = '#,##0.0000'
            ws.cell(i,16).number_format = '#,##0'
            i+=1

    #esto estamos probando a ver si sirve para que sea archivo descargable
                    
        buf=io.BytesIO()
        wb.save(buf)

        out=base64.b64encode(buf.getvalue())             
        buf.close()        

        #self.env.cr.execute((""" DELETE FROM vistaexcel_modelo"""))
        #self.invalidate_cache()

        modelo_excel = self.env['vistaexcel.modelo']
        nom_archivo = 'regventas_' + self.periodo_consulta + '.xlsx'
        move_vals = {
        'txt_filename': nom_archivo,
        'txt_binary': out,                                 
        }
        id_creado = modelo_excel.create(move_vals)
        
        return{
        'view_mode': 'form',
        'res_id': id_creado.id,
        'res_model': 'vistaexcel.modelo',
        'view_type': 'form',
        'type': 'ir.actions.act_window',
        'target': 'new',
        'flags': {'action_buttons': False},
        }

        
    #termina funcion que llama al reporte
'''class VistaExcelModel(models.Model):
    _name="vistaexcel.modelo"

    txt_filename = fields.Char()
    txt_binary = fields.Binary()
    id_excel= fields.Char()'''

   
    
